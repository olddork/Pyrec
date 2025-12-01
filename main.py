import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib.animation as animation
import matplotlib.dates as mdates
import math
import random
import csv
import time
import threading
import queue
from datetime import datetime, timedelta, date
import configparser
import os
import glob
import sys
from collections import deque
import itertools
import bisect  # Added for efficient searching

# --- DEPENDENCY CHECK ---
try:
    import serial
    import serial.tools.list_ports  
    SERIAL_AVAILABLE = True
except ImportError:
    SERIAL_AVAILABLE = False

try:
    import pandas as pd
    import openpyxl
    EXPORT_AVAILABLE = True
except ImportError:
    EXPORT_AVAILABLE = False

# --- Configuration ---
CONFIG_VERSION = 1  # Increment when INI schema changes
MAX_CHANNELS = 8
DEFAULT_WINDOW_SIZE = 50 
INI_FILE = "sensor_settings.ini"

# TIME CONSTANTS
SECONDS_PER_MINUTE = 60
SECONDS_PER_HOUR = 3600
SECONDS_PER_DAY = 86400

# DATA PROCESSING
GAP_MULTIPLIER = 20.0  # Increased to prevent line breakage when switching from slow to fast intervals
CHANNEL_NAME_TEMPLATE = "Ch_{}"  # Use .format(i+1) for channel names

# SLIDER RANGE (logarithmic scale)
SLIDER_MIN_SECONDS = 60       # 1 minute
SLIDER_MAX_SECONDS = 86400    # 24 hours

# COLORS & FONTS
COLOR_BG = "#f0f2f5"        
COLOR_SIDEBAR = "#ffffff"   
COLOR_ACCENT = "#007acc"    
COLOR_TEXT = "#2d3436"      
FONT_MAIN = ("Segoe UI", 10)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_HEADER = ("Segoe UI", 12, "bold")
FONT_MONO = ("Consolas", 10)

# PERFORMANCE & MEMORY
# Max points to RENDER on screen at once (Downsampling Target)
TARGET_PLOT_POINTS = 2000 
# Max points to STORE in RAM (approx 27 hours @ 1s interval)
MAX_BUFFER_SIZE = 100000 

INTERVAL_OPTIONS = {
    "1s": 1, "3s": 3, "5s": 5, "10s": 10, "30s": 30,
    "1min": 60, "2min": 120, "5min": 300, "10min": 600
}

BAUD_RATES = [9600, 19200, 38400, 57600, 115200, 250000]

# ==========================================
#        TOOLTIP CLASS
# ==========================================
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.widget.bind("<Enter>", self.show_tip)
        self.widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        try:
            x, y, cx, cy = self.widget.bbox("insert")
        except Exception:
            x, y, cx, cy = 0, 0, 0, 0
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 25
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(1)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hide_tip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

# ==========================================
#        DATA DRIVER ARCHITECTURE
# ==========================================

class DataSource:
    """Base class for all data sources (simulation, serial devices, etc.)."""
    
    def __init__(self) -> None:
        self.connected: bool = False
        self.latest_data: list[float] = [0.0] * MAX_CHANNELS
        self.lock: threading.Lock = threading.Lock()

    def connect(self, port: str, baud: int) -> bool:
        """Connect to the data source. Returns True on success."""
        self.connected = True
        return True

    def disconnect(self) -> None:
        """Disconnect from the data source."""
        self.connected = False

    def get_data(self) -> list[float]:
        """Return the latest sensor readings (thread-safe)."""
        with self.lock:
            return list(self.latest_data)

class SimulationSource(DataSource):
    def __init__(self):
        super().__init__()
        self.start_time = time.time()
        # Random offsets for each channel to make them look distinct
        self.phase_offsets = [random.uniform(0, 6.28) for _ in range(MAX_CHANNELS)]
        self.cloud_seeds = [random.uniform(0, 1000) for _ in range(MAX_CHANNELS)]

    def get_data(self) -> list[float]:
        t = time.time() - self.start_time
        data: list[float] = []
        
        for i in range(MAX_CHANNELS):
            # 1. Base "Sunlight" Curve (Very slow sine wave)
            # Period: 0.0005 rad/s ~= 3.5 hours for full cycle. Slow drift.
            sun_base = 2.5 + 2.0 * math.sin(t * 0.0005 + self.phase_offsets[i])
            
            # 2. "Cloud" Interference (Medium frequency subtractive noise)
            # Combine two faster sines to create irregular patterns
            c1 = math.sin(t * 0.02 + self.cloud_seeds[i])
            c2 = math.sin(t * 0.07 + self.cloud_seeds[i]*2)
            cloud_noise = (c1 + c2) / 2.0 # Range -1 to 1
            
            # If noise is high enough, it simulates a cloud blocking light (drop in value)
            drop = 0
            if cloud_noise > 0.4:
                drop = (cloud_noise - 0.4) * 2.0 # Magnitude of drop
            
            val = sun_base - drop
            
            # 3. Tiny Sensor Noise (Jitter)
            jitter = random.uniform(-0.01, 0.01)
            
            final_val = max(0.0, val + jitter) # Clip at 0
            data.append(final_val)
            
        return data

class GenericSerialDriver(DataSource):
    """Expects simple CSV line: 1.1, 2.2, 3.3..."""
    
    def __init__(self) -> None:
        super().__init__()
        self.ser: "serial.Serial | None" = None
        self.running: bool = False
        self.thread: threading.Thread | None = None

    def connect(self, port: str, baud: int) -> bool:
        if not SERIAL_AVAILABLE:
            messagebox.showerror("Error", "pyserial not installed.")
            return False
        try:
            self.ser = serial.Serial(port, baud, timeout=2)
            self.running = True
            self.thread = threading.Thread(target=self._reader_loop, daemon=True)
            self.thread.start()
            self.connected = True
            return True
        except Exception as e:
            messagebox.showerror("Connection Failed", str(e))
            return False

    def disconnect(self) -> None:
        self.running = False
        self.connected = False
        if self.ser and self.ser.is_open:
            self.ser.close()

    def _reader_loop(self) -> None:
        while self.running and self.ser and self.ser.is_open:
            try:
                line = self.ser.readline().decode('utf-8', errors='ignore').strip()
                if not line: continue
                
                parts = line.split(',')
                # Try to parse at least one value
                temp_data: list[float] = []
                for p in parts:
                    try: 
                        temp_data.append(float(p))
                    except ValueError: pass
                
                if temp_data:
                    # Pad with 0 if fewer than MAX_CHANNELS
                    while len(temp_data) < MAX_CHANNELS:
                        temp_data.append(0.0)
                    
                    with self.lock:
                        self.latest_data = temp_data[:MAX_CHANNELS]
            except Exception as e:
                print(f"GenericSerialDriver read error: {e}")
                time.sleep(0.1)

class BalkonLoggerDriver(DataSource):
    """Driver for BalkonLogger device using EOF-delimited batch protocol."""
    
    def __init__(self) -> None:
        super().__init__()
        self.ser: "serial.Serial | None" = None
        self.running: bool = False
        self.thread: threading.Thread | None = None
        self.synced: bool = False  # Track if we have found the first EOF

    def connect(self, port: str, baud: int) -> bool:
        if not SERIAL_AVAILABLE:
            messagebox.showerror("Error", "pyserial not installed.\nRun: pip install pyserial")
            return False
            
        try:
            # timeout=3 is important here to prevent blocking forever if device dies
            self.ser = serial.Serial(port, baud, timeout=3)
            self.running = True
            self.synced = False
            self.thread = threading.Thread(target=self._reader_loop, daemon=True)
            self.thread.start()
            self.connected = True
            return True
        except Exception as e:
            messagebox.showerror("Connection Failed", str(e))
            return False

    def disconnect(self) -> None:
        self.running = False
        self.connected = False
        if self.ser and self.ser.is_open:
            self.ser.close()

    def _reader_loop(self) -> None:
        # Debug: print("Serial thread started...")
        while self.running and self.ser and self.ser.is_open:
            try:
                # Read a line
                line_bytes = self.ser.readline()
                line = line_bytes.decode('utf-8', errors='ignore').strip()
                
                if not line: continue

                # --- SYNC LOGIC ---
                # We do not record any data until we hit the first "eof" marker.
                # This ensures we don't assign Channel 8 value to Channel 1 by accident.
                if not self.synced:
                    if line == "eof":
                        self.synced = True
                        # Debug: print("Sync achieved. Waiting for data batch...")
                    continue

                # --- BATCH READ LOGIC ---
                # If we are here, the LAST line seen was "eof". 
                # We expect the NEXT 16 lines to be our data block.
                if line == "eof":
                    temp_data = []
                    
                    # We need to read 16 lines. 
                    # We use a loop with a hard break to prevent getting stuck.
                    for _ in range(16):
                        val_str = self.ser.readline().decode('utf-8', errors='ignore').strip()
                        
                        # Safety: If we see "eof" again inside the block, we lost sync.
                        if val_str == "eof":
                            temp_data = [] # Discard corrupt batch
                            break
                            
                        try:
                            # Parse float
                            val = float(val_str)
                            temp_data.append(val)
                        except ValueError:
                            pass 
                    
                    # BalkonLogger specific: It sends 16 values, usually only first 8 are used
                    if len(temp_data) >= 8:
                        with self.lock:
                            self.latest_data = temp_data[:8]
                    else:
                        pass  # Debug: print("Incomplete batch received. Discarding.")
                        
            except Exception as e:
                # Keep error logging for debugging serial issues
                print(f"Serial Error: {e}")
                self.synced = False # Force re-sync on error
                time.sleep(1)

DEVICE_DRIVERS = {
    "BalkonLogger": BalkonLoggerDriver,
    "Standard": GenericSerialDriver  # Added for simple 1s interval support
}

# ==========================================
#        EXPORT HANDLER
# ==========================================
class ExportHandler:
    @staticmethod
    def show_dialog(parent, x_limits, ch_vars):
        if not EXPORT_AVAILABLE:
             messagebox.showerror("Error", "pandas/openpyxl missing. Install with pip.")
             return

        top = tk.Toplevel(parent)
        top.title("Export View Data")
        top.geometry("500x500") 
        top.configure(bg=COLOR_BG)
        c_frame = ttk.Frame(top, padding=20)
        c_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(c_frame, text="Export Settings", font=FONT_HEADER).pack(pady=(0,10))
        
        ch_frame = ttk.LabelFrame(c_frame, text="Select Channels to Export", padding=10)
        ch_frame.pack(fill=tk.X, pady=10)
        
        export_vars = []
        for i in range(MAX_CHANNELS):
            is_active = ch_vars['active'][i].get()
            var = tk.BooleanVar(value=is_active)
            export_vars.append(var)
            col = i % 2
            row = i // 2
            ttk.Checkbutton(ch_frame, text=f"Channel {i+1}", variable=var).grid(row=row, column=col, sticky="w", padx=10, pady=2)

        x_min, x_max = x_limits
        try:
            start_dt = mdates.num2date(x_min).replace(tzinfo=None) 
            end_dt = mdates.num2date(x_max).replace(tzinfo=None)
        except:
            start_dt = datetime.now() - timedelta(hours=1)
            end_dt = datetime.now()

        ttk.Label(c_frame, text="Start Time (YYYY-MM-DD HH:MM):").pack(anchor="w")
        s_entry = ttk.Entry(c_frame, width=30)
        s_entry.insert(0, start_dt.strftime("%Y-%m-%d %H:%M"))
        s_entry.pack(fill=tk.X, pady=(5, 5))
        
        ttk.Label(c_frame, text="End Time (YYYY-MM-DD HH:MM):").pack(anchor="w")
        e_entry = ttk.Entry(c_frame, width=30)
        e_entry.insert(0, end_dt.strftime("%Y-%m-%d %H:%M"))
        e_entry.pack(fill=tk.X, pady=(5, 20))
        
        def run_export_thread():
            try:
                s_d = datetime.strptime(s_entry.get(), "%Y-%m-%d %H:%M")
                e_d = datetime.strptime(e_entry.get(), "%Y-%m-%d %H:%M")
                if s_d >= e_d:
                    messagebox.showerror("Error", "Start time must be before end time.")
                    return
                
                sel_chs = [v.get() for v in export_vars]
                
                # Capture current channel settings (Factors, Offsets, Colors) to pass to thread safely
                factors = [ch_vars['factor'][i].get() for i in range(MAX_CHANNELS)]
                offsets = [ch_vars['offset'][i].get() for i in range(MAX_CHANNELS)]
                colors = [ch_vars['colors'][i] for i in range(MAX_CHANNELS)]
                
                out = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile=f"export.xlsx")
                if not out: return

                # UI Feedback
                btn_go.config(text="Processing... (Please Wait)", state="disabled")
                
                # Start Thread
                threading.Thread(target=ExportHandler.process_thread, args=(s_d, e_d, sel_chs, factors, offsets, colors, out, top, parent), daemon=True).start()
                
            except ValueError: 
                messagebox.showerror("Format Error", "Invalid Date Format.\nUse: YYYY-MM-DD HH:MM")

        btn_go = ttk.Button(c_frame, text="Export Now", command=run_export_thread)
        btn_go.pack(fill=tk.X, pady=10)

    @staticmethod
    def process_thread(start_dt, end_dt, channels_to_export, factors, offsets, colors, out_path, popup, parent):
        """Heavy lifting for export in a separate thread."""
        try: 
            import pandas as pd
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill, Font
            
            all_logs = glob.glob("log_*.csv")
            relevant_files = []
            req_start_date = start_dt.date()
            req_end_date = end_dt.date()

            for log in all_logs:
                try:
                    date_str = log.replace("log_", "").replace(".csv", "")
                    file_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                    if req_start_date <= file_date <= req_end_date:
                        relevant_files.append(log)
                except: continue 
            relevant_files.sort()
            
            if not relevant_files:
                parent.after(0, lambda: messagebox.showwarning("Empty", "No log files found."))
                parent.after(0, popup.destroy)
                return

            combined = []
            start_ts = start_dt.timestamp()
            end_ts = end_dt.timestamp()

            for log in relevant_files:
                try:
                    chunk_iter = pd.read_csv(log, comment='#', on_bad_lines='skip', chunksize=10000)
                    for chunk in chunk_iter:
                        if 'Timestamp_Unix' not in chunk.columns: continue
                        mask = (chunk['Timestamp_Unix'] >= start_ts) & (chunk['Timestamp_Unix'] <= end_ts)
                        filtered = chunk.loc[mask]
                        if not filtered.empty:
                            if 'Timestamp_ISO' in filtered.columns: 
                                final = filtered.drop(columns=['Timestamp_Unix'])
                            else:
                                filtered['Timestamp'] = pd.to_datetime(filtered['Timestamp_Unix'], unit='s')
                                final = filtered.drop(columns=['Timestamp_Unix'])
                            combined.append(final)
                except Exception as e: 
                    print(f"Skipping bad file/chunk: {e}")

            if combined:
                full_df = pd.concat(combined, ignore_index=True)
                rename_map = {CHANNEL_NAME_TEMPLATE.format(i+1): f"Raw_{CHANNEL_NAME_TEMPLATE.format(i+1)}" for i in range(MAX_CHANNELS)}
                full_df.rename(columns=rename_map, inplace=True)
                
                # Filter cols
                for i in range(MAX_CHANNELS):
                    if not channels_to_export[i]:
                        col_name = f"Raw_{CHANNEL_NAME_TEMPLATE.format(i+1)}"
                        if col_name in full_df.columns: full_df.drop(columns=[col_name], inplace=True)

                # Add calc placeholders
                for i in range(MAX_CHANNELS):
                    if channels_to_export[i]: full_df[f"Cal_{CHANNEL_NAME_TEMPLATE.format(i+1)}"] = "" 

                # Write Excel
                with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                    full_df.to_excel(writer, index=False, startrow=3)
                
                # Formatting
                wb = openpyxl.load_workbook(out_path)
                ws = wb.active
                ws['A2'] = "Factor"; ws['A3'] = "Offset"
                ws['A2'].font = Font(bold=True); ws['A3'].font = Font(bold=True)

                header_row = 4
                col_map = {}
                for cell in ws[header_row]:
                    if cell.value:
                        col_map[cell.value] = get_column_letter(cell.column)
                        ws.column_dimensions[get_column_letter(cell.column)].width = 18 
                
                max_row = ws.max_row
                
                # Formula Injection
                for i in range(MAX_CHANNELS):
                    if not channels_to_export[i]: continue
                    raw_col = f"Raw_{CHANNEL_NAME_TEMPLATE.format(i+1)}"
                    cal_col = f"Cal_{CHANNEL_NAME_TEMPLATE.format(i+1)}"
                    
                    if raw_col in col_map and cal_col in col_map:
                        r_let = col_map[raw_col]
                        c_let = col_map[cal_col]
                        
                        f_val = factors[i]
                        o_val = offsets[i]
                        
                        ws[f"{c_let}2"] = f_val
                        ws[f"{c_let}3"] = o_val
                        
                        for row in range(5, max_row + 1):
                            ws[f"{c_let}{row}"] = f"={r_let}{row}*{c_let}$2+{c_let}$3"
                            
                        hex_color = colors[i].replace('#', '')
                        fill = PatternFill(start_color="FF"+hex_color, end_color="FF"+hex_color, fill_type="solid")
                        ws[f"{c_let}{header_row}"].fill = fill

                wb.save(out_path)
                parent.after(0, lambda: messagebox.showinfo("Success", f"Exported to {os.path.basename(out_path)}"))
                parent.after(0, popup.destroy)
            else: 
                parent.after(0, lambda: messagebox.showwarning("Empty", "No data in range."))
                parent.after(0, popup.destroy)
                
        except Exception as e: 
            parent.after(0, lambda: messagebox.showerror("Export Error", str(e)))
            parent.after(0, popup.destroy)

# ==========================================
#              MAIN APPLICATION
# ==========================================

class CustomToolbar(NavigationToolbar2Tk):
    def __init__(self, canvas, window, app_instance):
        self.app = app_instance
        super().__init__(canvas, window)
        self.config(background=COLOR_BG)
        
        # 1. HIDE STANDARD BUTTONS
        for child in self.winfo_children():
            child.pack_forget()

        # 2. CREATE CUSTOM CONTAINERS
        self.frame_left = tk.Frame(self, bg=COLOR_BG)
        self.frame_left.pack(side=tk.LEFT, padx=2)
        self.frame_right = tk.Frame(self, bg=COLOR_BG)
        self.frame_right.pack(side=tk.RIGHT, padx=2)

        # ---------------- LEFT GROUP ----------------
        self.port_var = tk.StringVar(value="Simulation")
        self.cb_port = ttk.Combobox(self.frame_left, textvariable=self.port_var, values=["Simulation"], 
                                    state="readonly", width=15, font=FONT_MAIN, 
                                    postcommand=self.refresh_ports)
        self.cb_port.pack(side=tk.LEFT, padx=2)
        
        self.baud_var = tk.StringVar(value="9600")
        self.cb_baud = ttk.Combobox(self.frame_left, textvariable=self.baud_var, values=BAUD_RATES, 
                                    state="readonly", width=7, font=FONT_MAIN)
        self.cb_baud.pack(side=tk.LEFT, padx=2)

        self.device_var = tk.StringVar(value="Standard")
        self.cb_device = ttk.Combobox(self.frame_left, textvariable=self.device_var, values=list(DEVICE_DRIVERS.keys()), 
                                      state="readonly", width=12, font=FONT_MAIN)
        self.cb_device.pack(side=tk.LEFT, padx=2)
        self.cb_device.bind("<<ComboboxSelected>>", self.app.on_device_changed)

        self.btn_connect = tk.Button(self.frame_left, text="🔌 Connect", command=self.toggle_connect,
                                     bg="#e74c3c", fg="white", font=FONT_BOLD, relief="flat", width=12,
                                     activebackground="#c0392b", activeforeground="white", cursor="hand2")
        self.btn_connect.pack(side=tk.LEFT, padx=5)

        # ---------------- RIGHT GROUP ----------------
        # Modern flat button style with hover effects
        self.btn_style_normal = {
            "relief": "flat", 
            "bg": "#ffffff", 
            "fg": "#2d3436",
            "font": ("Segoe UI", 13), 
            "cursor": "hand2", 
            "width": 3, 
            "height": 1,
            "activebackground": "#dfe6e9",
            "activeforeground": "#2d3436",
            "bd": 0,
            "highlightthickness": 1,
            "highlightbackground": "#dfe6e9",
            "highlightcolor": COLOR_ACCENT,
            "takefocus": False 
        }

        self.lbl_lock_status = tk.Label(self.frame_right, text="🔒", font=("Segoe UI", 16), bg=COLOR_BG, fg="#27ae60")
        self.lbl_lock_status.pack(side=tk.LEFT, padx=10)
        ToolTip(self.lbl_lock_status, "Auto-Scroll: Active")

        # Define buttons structure: (AttributeName, Label, Command, Tooltip)
        buttons_config = [
            ("btn_home", "🏠", self.home, "Reset View (Home)"),
            ("btn_pan", "✋", self.pan, "Pan Mode"),
            ("btn_zoom", "🔍", self.zoom, "Zoom Mode"),
            ("btn_save", "💾", self.save_figure, "Save Screenshot"),
            ("btn_pause", "⏸", self.toggle_pause, "Pause/Resume"),
            ("btn_export", "📊", self.app.open_export_window, "Export Data")
        ]

        # Generate buttons in loop and assign to self
        for attr, text, cmd, tooltip in buttons_config:
            btn = tk.Button(self.frame_right, text=text, command=cmd, **self.btn_style_normal)
            btn.pack(side=tk.LEFT, padx=3, pady=4)
            ToolTip(btn, tooltip)
            setattr(self, attr, btn) # Store explicitly so other methods can access self.btn_pan etc.

    def home(self, *args):
        # Reset Y-axis to UI box values, keep current slider, enable live mode
        self.app.apply_scale()  # Reset Y to box values
        self.app.auto_scroll.set(True)  # Enable live mode (X will snap to current time)
        # Reset toolbar mode (deselect pan/zoom)
        if self.mode:
            if 'pan' in str(self.mode).lower():
                self.pan()  # Toggle off
            elif 'zoom' in str(self.mode).lower():
                self.zoom()  # Toggle off
        self._update_buttons_state()

    def zoom(self):
        super().zoom()
        self._update_buttons_state()

    def pan(self):
        super().pan()
        self._update_buttons_state()

    def _update_buttons_state(self):
        mode = str(self.mode).lower()
        # Active button style (selected)
        active_style = {"bg": COLOR_ACCENT, "fg": "white", "highlightbackground": COLOR_ACCENT}
        # Normal button style (not selected)
        normal_style = {"bg": "#ffffff", "fg": "#2d3436", "highlightbackground": "#dfe6e9"}
        
        if 'pan' in mode:
            self.app.auto_scroll.set(False)
            self.lbl_lock_status.config(text="🔓", fg="#e74c3c")
            self.btn_pan.config(**active_style)
            self.btn_zoom.config(**normal_style)
        elif 'zoom' in mode:
            self.app.auto_scroll.set(False)
            self.lbl_lock_status.config(text="🔓", fg="#e74c3c")
            self.btn_pan.config(**normal_style)
            self.btn_zoom.config(**active_style)
        else:
            # Transitioning to live mode - capture current X view to slider
            self.app.capture_current_x_view()
            self.app.auto_scroll.set(True)
            self.lbl_lock_status.config(text="🔒", fg="#27ae60")
            self.btn_pan.config(**normal_style)
            self.btn_zoom.config(**normal_style)

    def toggle_pause(self):
        self.app.toggle_capture()
        if self.app.is_running: 
            self.btn_pause.config(text="⏸", fg="#2d3436", bg="#ffffff", highlightbackground="#dfe6e9")
        else: 
            self.btn_pause.config(text="▶", fg="white", bg="#e74c3c", highlightbackground="#e74c3c") 

    def refresh_ports(self):
        values = ["Simulation"]
        if SERIAL_AVAILABLE:
            try:
                ports = serial.tools.list_ports.comports()
                for p in ports: values.append(p.device)
            except: pass
        self.cb_port['values'] = values

    def toggle_connect(self):
        if not self.app.connected:
            port = self.port_var.get()
            baud = int(self.baud_var.get())
            device_name = self.device_var.get()
            
            if self.app.connect_to_source(port, baud, device_name):
                self.btn_connect.config(text="✔ Connected", bg="#27ae60", activebackground="#1e8449")
                self.cb_port.config(state="disabled")
                self.cb_baud.config(state="disabled")
                self.cb_device.config(state="disabled")
        else:
            self.app.disconnect_source()
            self.btn_connect.config(text="🔌 Connect", bg="#e74c3c", activebackground="#c0392b")
            self.cb_port.config(state="readonly")
            self.cb_baud.config(state="readonly")
            self.cb_device.config(state="readonly")

class SensorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pyrec Data 0.9 beta")
        self.root.geometry("1280x850")
        self.root.configure(bg=COLOR_BG)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        # --- Data Source ---
        self.data_source = None
        self.connected = False
        
        # --- Data Storage ---
        self.timestamps = deque(maxlen=MAX_BUFFER_SIZE)
        self.datetime_cache = deque(maxlen=MAX_BUFFER_SIZE) 
        self.channel_data = [deque(maxlen=MAX_BUFFER_SIZE) for _ in range(MAX_CHANNELS)]
        
        # --- PERFORMANCE OPTIMIZATION ---
        # Cache the timestamps list to avoid O(N) conversion on every frame
        self._timestamps_list_cache = []
        self._timestamps_cache_size = 0
        self._datetime_axis_initialized = False
        
        # --- THREADED LOGGING SETUP ---
        self.log_queue = queue.Queue()
        self.file_lock = threading.RLock() 
        self.log_thread = threading.Thread(target=self._log_worker, daemon=True)
        self.log_thread.start()
        
        # --- State ---
        self.is_running = True
        self.start_time = time.time()
        self.last_capture_time = 0 
        self.current_log_date = None
        self.log_file = None
        
        # --- Slider Log Scale Parameters (use constants from file top) ---
        self.slider_min_log = math.log(SLIDER_MIN_SECONDS)
        self.slider_max_log = math.log(SLIDER_MAX_SECONDS)
        self.slider_log_range = self.slider_max_log - self.slider_min_log
        self.csv_writer = None
        self.log_filename = ""
        
        # --- UI Vars ---
        self.auto_scroll = tk.BooleanVar(value=True)
        self.window_size_var = tk.IntVar(value=56)  # Slider 0-100 scale (log), 56 ≈ 1 hour
        self.y_min = tk.DoubleVar(value=-0.5) 
        self.y_max = tk.DoubleVar(value=4.5)  
        self.save_on_exit_var = tk.BooleanVar(value=True)
        self.interval_var = tk.StringVar(value="1s") 
        self.current_interval_sec = 1.0 
        
        self.ch_vars = {
            'active': [tk.BooleanVar(value=True if i < 3 else False) for i in range(MAX_CHANNELS)],
            'factor': [tk.DoubleVar(value=1.0) for _ in range(MAX_CHANNELS)],
            'offset': [tk.DoubleVar(value=0.0) for _ in range(MAX_CHANNELS)],
            'colors': ['#2980b9', '#27ae60', '#c0392b', '#16a085', '#8e44ad', '#f39c12', '#2c3e50', '#d35400'],
            'current_val_str': [tk.StringVar(value="0.00") for _ in range(MAX_CHANNELS)]
        }

        self._apply_theme()
        self.load_settings()
        self.init_daily_log_system() # Setup file immediately
        
        # --- UI Setup ---
        self._setup_ui()
        self.update_slider_range()
        self.on_device_changed(None) # Trigger rule check for device
        self.apply_scale() # <--- Force application of Y-Axis limits from settings
        self.on_slider_drag(self.window_size_var.get())  # Initialize slider label display

        # --- BACKGROUND LOADING (HISTORY) ---
        # We start a thread to load old data so the UI doesn't freeze on startup
        threading.Thread(target=self._load_history_worker, daemon=True).start()
        
        # Auto-connect Simulation
        self.connect_to_source("Simulation", 9600, "Standard")
        self.toolbar_ref.btn_connect.config(text="✔ Disconnect", bg="#27ae60")
        self.toolbar_ref.cb_port.config(state="disabled")
        
        # --- ANIMATION ---
        self.ani = animation.FuncAnimation(self.fig, self.update_plot, init_func=self.init_plot, interval=200, blit=False, cache_frame_data=False)

    def _apply_theme(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure(".", background=COLOR_BG, foreground=COLOR_TEXT, font=FONT_MAIN)
        style.configure("TLabel", background=COLOR_BG, font=FONT_MAIN)
        style.configure("TButton", font=FONT_BOLD, background="#dfe6e9", borderwidth=0)
        style.map("TButton", background=[("active", "#b2bec3")])
        style.configure("Card.TFrame", background=COLOR_SIDEBAR, relief="flat")
        style.configure("Card.TLabel", background=COLOR_SIDEBAR, font=FONT_MAIN)
        style.configure("Header.TLabel", background=COLOR_SIDEBAR, font=FONT_HEADER, foreground=COLOR_ACCENT)
        style.configure("Status.TLabel", background=COLOR_SIDEBAR, foreground="#27ae60", font=("Consolas", 8))

    def connect_to_source(self, port: str, baud: int, device_name: str) -> bool:
        """Connect to a data source (Simulation or serial device)."""
        if port == "Simulation":
            DriverClass = SimulationSource
        else:
            DriverClass = DEVICE_DRIVERS.get(device_name, BalkonLoggerDriver)
            
        self.data_source = DriverClass()
        if self.data_source.connect(port, baud):
            self.connected = True
            self.lbl_status.config(text=f"Connected: {port}\nDevice: {device_name}")
            return True
        else:
            self.data_source = None
            self.connected = False
            return False

    def disconnect_source(self) -> None:
        """Disconnect from the current data source."""
        if self.data_source:
            self.data_source.disconnect()
        self.data_source = None
        self.connected = False
        self.lbl_status.config(text=f"Status: Disconnected\nLog: {self.log_filename}")

    # --- LOGGING WORKER (Background Thread) ---
    def _log_worker(self) -> None:
        """Background thread that writes log rows from queue to CSV file."""
        while True:
            item = self.log_queue.get()
            if item is None:
                self.log_queue.task_done()
                break
            
            with self.file_lock:
                if self.csv_writer and self.log_file:
                    try:
                        self.csv_writer.writerow(item)
                    except Exception as e:
                        print(f"Write Error: {e}")
            
            self.log_queue.task_done()

    def get_daily_filename(self):
        return f"log_{datetime.now().strftime('%Y-%m-%d')}.csv"

    def init_daily_log_system(self):
        with self.file_lock:
            now = datetime.now()
            today_str = now.strftime('%Y-%m-%d')
            self.current_log_date = today_str
            self.log_filename = self.get_daily_filename()

            try:
                if self.log_file and not self.log_file.closed:
                    self.log_file.close()

                file_exists = os.path.exists(self.log_filename)
                self.log_file = open(self.log_filename, mode='a', newline='')
                self.csv_writer = csv.writer(self.log_file)
                if not file_exists:
                    self.csv_writer.writerow(["# DAILY LOG START", today_str])
                    self.csv_writer.writerow(["Timestamp_Unix", "Timestamp_ISO"] + [CHANNEL_NAME_TEMPLATE.format(i+1) for i in range(MAX_CHANNELS)])
            except Exception as e:
                messagebox.showerror("File Error", str(e))
                self.is_running = False

    # --- HISTORY LOADING WORKER ---
    def _load_history_worker(self):
        """Loads data from CSV files in background without blocking UI."""
        # Note: We are modifying self.timestamps (a deque) from a thread. 
        # Deque append is atomic in CPython (mostly safe), but for 100% safety we usually use a lock.
        # Given the app startup sequence, this runs before live data starts pouring in rapidly.
        
        all_logs = glob.glob("log_*.csv")
        all_logs.sort()
        
        # Only load last 3 days
        relevant_files = []
        cutoff_date = (datetime.now() - timedelta(days=3)).date() 
        
        for fname in all_logs:
            try:
                d_str = os.path.basename(fname).replace("log_", "").replace(".csv", "")
                f_date = datetime.strptime(d_str, "%Y-%m-%d").date()
                if f_date >= cutoff_date:
                    relevant_files.append(fname)
            except: pass 
        
        # Debug: print(f"Loading history from {len(relevant_files)} files...")

        # Temp buffers to minimize lock contention
        temp_ts = []
        temp_dt = []
        temp_ch = [[] for _ in range(MAX_CHANNELS)]

        prev_ts = None
        for fname in relevant_files:
            try:
                with open(fname, 'r') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if not row or row[0].startswith("#") or row[0].startswith("Timestamp"): continue
                        try:
                            ts_val = float(row[0])
                            
                            # Gap detection (add a NaN point if gap > 5x interval)
                            if prev_ts is not None and (ts_val - prev_ts) > (self.current_interval_sec * 5):
                                pad_ts = ts_val - 0.001
                                temp_ts.append(pad_ts)
                                temp_dt.append(datetime.fromtimestamp(pad_ts))
                                for i in range(MAX_CHANNELS): temp_ch[i].append(float('nan'))
                            
                            temp_ts.append(ts_val)
                            temp_dt.append(datetime.fromtimestamp(ts_val))
                            
                            for i in range(MAX_CHANNELS):
                                col_idx = 2 + i
                                val = float(row[col_idx]) if col_idx < len(row) else 0.0
                                temp_ch[i].append(val)
                            prev_ts = ts_val
                        except ValueError: continue
            except Exception as e:
                # Keep error logging for debugging file issues
                print(f"Error reading {fname}: {e}")

        # Bulk update main buffers (quick operation)
        if temp_ts:
            # We use extend to push to deque
            self.timestamps.extend(temp_ts)
            self.datetime_cache.extend(temp_dt)
            for i in range(MAX_CHANNELS):
                self.channel_data[i].extend(temp_ch[i])
            
            self.last_capture_time = temp_ts[-1]
            # Debug: print(f"History loaded: {len(temp_ts)} points.")

    def check_rollover(self):
        if datetime.now().strftime('%Y-%m-%d') != self.current_log_date:
            with self.file_lock: 
                if self.log_file: self.log_file.close()
                self.init_daily_log_system()
                if hasattr(self, 'lbl_status'):
                    self.lbl_status.config(text=f"Active Log:\n{self.log_filename}")

    # --- UI LAYOUT ---
    def _setup_ui(self):
        sidebar = ttk.Frame(self.root, padding="15", width=350, style="Card.TFrame")
        sidebar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.lbl_status = ttk.Label(sidebar, text=f"Source: Simulation\nLog: {self.log_filename}", style="Status.TLabel")
        self.lbl_status.pack(pady=(0, 15), anchor="w")
        
        ttk.Label(sidebar, text="Sampling Rate", style="Header.TLabel").pack(anchor="w", pady=(0, 5))
        self.interval_cb = ttk.Combobox(sidebar, textvariable=self.interval_var, values=list(INTERVAL_OPTIONS.keys()), state="readonly", font=FONT_MAIN)
        self.interval_cb.pack(fill=tk.X, pady=(0, 15))
        self.interval_cb.bind("<<ComboboxSelected>>", self.on_interval_change)
        
        ttk.Label(sidebar, text="Channels", style="Header.TLabel").pack(anchor="w", pady=(0, 5))
        grid_frame = ttk.Frame(sidebar, style="Card.TFrame")
        grid_frame.pack(fill=tk.X)
        headers = ["Ch", "On", "Value", "Factor", "Offset"]
        for col, txt in enumerate(headers):
            ttk.Label(grid_frame, text=txt, style="Card.TLabel", font=("Segoe UI", 8, "bold")).grid(row=0, column=col, sticky="w", padx=2)

        for i in range(MAX_CHANNELS):
            row = i + 1
            lbl = ttk.Label(grid_frame, text=f"■ {i+1}", foreground=self.ch_vars['colors'][i], font=FONT_BOLD, style="Card.TLabel")
            lbl.grid(row=row, column=0, padx=2, pady=4, sticky="w")
            cb = ttk.Checkbutton(grid_frame, variable=self.ch_vars['active'][i], style="Card.TCheckbutton")
            cb.grid(row=row, column=1)
            v_lbl = ttk.Label(grid_frame, textvariable=self.ch_vars['current_val_str'][i], width=8, anchor="e", font=FONT_MONO, background="#f1f2f6")
            v_lbl.grid(row=row, column=2, padx=5)
            
            spin_f = ttk.Spinbox(grid_frame, from_=-1000.0, to=1000.0, increment=0.1, 
                                 textvariable=self.ch_vars['factor'][i], width=5, font=FONT_MONO)
            spin_f.grid(row=row, column=3, padx=2)
            
            spin_o = ttk.Spinbox(grid_frame, from_=-10000.0, to=10000.0, increment=0.1, 
                                 textvariable=self.ch_vars['offset'][i], width=5, font=FONT_MONO)
            spin_o.grid(row=row, column=4, padx=2)

        ttk.Separator(sidebar, orient='horizontal').pack(fill='x', pady=20)
        ttk.Label(sidebar, text="Display Settings", style="Header.TLabel").pack(anchor="w", pady=(0, 5))
        ttk.Label(sidebar, text="Time Window (Logarithmic)", style="Card.TLabel").pack(anchor="w")
        
        # SLIDER: 0-100 linear scale, converted to log time internally (60s to 86400s)
        self.time_slider = ttk.Scale(sidebar, from_=0, to=100, variable=self.window_size_var, orient='horizontal', command=self.on_slider_drag)
        self.time_slider.pack(fill=tk.X, pady=5)
        
        self.lbl_window_display = ttk.Label(sidebar, text="1h 0m", style="Card.TLabel")
        self.lbl_window_display.pack(anchor="w", pady=(0, 10))
        
        y_frame = ttk.Frame(sidebar, style="Card.TFrame")
        y_frame.pack(fill=tk.X, pady=5)
        ttk.Label(y_frame, text="Y-Min:", style="Card.TLabel").pack(side=tk.LEFT)
        e1 = ttk.Entry(y_frame, textvariable=self.y_min, width=6, font=FONT_MONO)
        e1.pack(side=tk.LEFT, padx=5)
        e1.bind('<Return>', lambda e: self.apply_scale())
        ttk.Label(y_frame, text="Y-Max:", style="Card.TLabel").pack(side=tk.LEFT)
        e2 = ttk.Entry(y_frame, textvariable=self.y_max, width=6, font=FONT_MONO)
        e2.pack(side=tk.LEFT, padx=5)
        e2.bind('<Return>', lambda e: self.apply_scale())
        ttk.Button(sidebar, text="Update Scale", command=self.apply_scale, takefocus=False).pack(fill=tk.X, pady=2)
        
        footer = ttk.Frame(sidebar, style="Card.TFrame")
        footer.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Checkbutton(footer, text="Save Settings on Exit", variable=self.save_on_exit_var, style="Card.TCheckbutton").pack(anchor="w", pady=(0, 10))
        
        tk.Button(footer, text="EXIT APPLICATION", command=self.on_close,
                  bg="#ffcccc", fg="red", font=("Segoe UI", 10, "bold"), relief="raised", takefocus=False).pack(fill=tk.X, pady=5)

        main_frame = ttk.Frame(self.root)
        main_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Use tight_layout=False and set fixed margins to prevent canvas shifting
        self.fig, self.ax = plt.subplots(figsize=(5, 4), dpi=100)
        self.fig.subplots_adjust(left=0.08, right=0.98, top=0.92, bottom=0.12)  # Fixed margins
        self.fig.patch.set_facecolor(COLOR_BG) 
        self.ax.set_facecolor("white")
        self.ax.set_title("Live Sensor Data", fontsize=12, fontweight='bold', color=COLOR_TEXT)
        self.ax.tick_params(colors=COLOR_TEXT, labelsize=9)
        self.ax.grid(True, color='#dfe6e9', linestyle='--', linewidth=0.5)
        self.ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M')) # Shortened format for long views
        
        self.lines = []
        for i in range(MAX_CHANNELS):
            line, = self.ax.plot([], [], label=f"Ch {i+1}", color=self.ch_vars['colors'][i], linewidth=1.5)
            self.lines.append(line)
        self.canvas = FigureCanvasTkAgg(self.fig, master=main_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        self.toolbar_ref = CustomToolbar(self.canvas, main_frame, self)
        self.toolbar_ref.update()
        self.toolbar_ref.pack(side=tk.BOTTOM, fill=tk.X) 

    def update_slider_range(self):
        # No longer needed - slider range is fixed to time (60s to 86400s)
        pass

    def slider_to_seconds(self, slider_val: float) -> float:
        """Convert slider value (0-100) to time window in seconds using log scale."""
        slider_fraction = float(slider_val) / 100.0
        log_time = self.slider_min_log + (slider_fraction * self.slider_log_range)
        return math.exp(log_time)

    def seconds_to_slider(self, seconds: float) -> int:
        """Convert time window in seconds to slider value (0-100) using log scale."""
        # Clamp to valid range
        seconds = max(SLIDER_MIN_SECONDS, min(SLIDER_MAX_SECONDS, seconds))
        log_time = math.log(seconds)
        slider_fraction = (log_time - self.slider_min_log) / self.slider_log_range
        return int(round(slider_fraction * 100.0))

    def capture_current_x_view(self) -> None:
        """Capture current X-axis time span and update slider to match."""
        try:
            x_min, x_max = self.ax.get_xlim()
            # Convert matplotlib date numbers to seconds
            view_min_dt = mdates.num2date(x_min).replace(tzinfo=None)
            view_max_dt = mdates.num2date(x_max).replace(tzinfo=None)
            time_span_seconds = (view_max_dt - view_min_dt).total_seconds()
            
            # Update slider to match current view
            new_slider_val = self.seconds_to_slider(time_span_seconds)
            self.window_size_var.set(new_slider_val)
            self.lbl_window_display.config(text=self.format_time_window(time_span_seconds))
        except Exception:
            pass  # Keep current slider value if conversion fails

    def format_time_window(self, seconds: float) -> str:
        """Format seconds as 'Xh Ym' string for display."""
        hours = int(seconds // SECONDS_PER_HOUR)
        mins = int((seconds % SECONDS_PER_HOUR) // SECONDS_PER_MINUTE)
        return f"{hours}h {mins}m"

    def on_slider_drag(self, val):
        """
        Handles slider interaction with logarithmic scaling.
        - Slider 0-100 (linear) converts to 60s-86400s (log)
        - If Auto-Scroll is ON: Just updates variable
        - If Auto-Scroll is OFF (Pan Mode): Manually expands/contracts view around CENTER
        """
        window_sec = self.slider_to_seconds(val)
        self.lbl_window_display.config(text=self.format_time_window(window_sec))
        
        if self.auto_scroll.get():
            # Force immediate update of the view without waiting for animation tick
            if self.timestamps:
                self._render_plot()
                self.canvas.draw_idle()
            return

        # PAN/ZOOM MODE LOGIC
        try:
            width_days = window_sec / SECONDS_PER_DAY
            
            # Get Current View Center
            current_min, current_max = self.ax.get_xlim()
            center = (current_min + current_max) / 2.0
            
            # Apply New Limits centered on existing view
            new_min = center - (width_days / 2.0)
            new_max = center + (width_days / 2.0)
            
            self.ax.set_xlim(new_min, new_max)
            self.canvas.draw_idle()
            
        except Exception:
            pass

    def on_device_changed(self, event):
        """Rule: Disable 1s interval for BalkonLogger"""
        current_dev = self.toolbar_ref.device_var.get()
        current_interval = self.interval_var.get()
        
        # Modify the values list in the combobox
        all_vals = list(INTERVAL_OPTIONS.keys())
        
        if current_dev == "BalkonLogger":
            # Remove '1s' from options
            if '1s' in all_vals: all_vals.remove('1s')
            
            # If current selection is 1s, force change to 3s
            if current_interval == '1s':
                self.interval_var.set('3s')
                self.current_interval_sec = 3
                
            self.interval_cb['values'] = all_vals
        else:
            # Restore all options
            self.interval_cb['values'] = list(INTERVAL_OPTIONS.keys())

    def on_interval_change(self, event):
        text_val = self.interval_var.get()
        if text_val in INTERVAL_OPTIONS:
            self.current_interval_sec = INTERVAL_OPTIONS[text_val]
            self.last_capture_time = 0 

    # --- EXPORT LOGIC (THREADED) ---
    def open_export_window(self):
        ExportHandler.show_dialog(self.root, self.ax.get_xlim(), self.ch_vars)

    def load_settings(self):
        if not os.path.exists(INI_FILE): return
        config = configparser.ConfigParser()
        try:
            config.read(INI_FILE)
        except Exception as e:
            print(f"Error reading {INI_FILE}: {e}")
            return
        try:
            if 'Graph' in config:
                self.window_size_var.set(config['Graph'].getint('window_size', DEFAULT_WINDOW_SIZE))
                self.y_min.set(config['Graph'].getfloat('y_min', -0.5))
                self.y_max.set(config['Graph'].getfloat('y_max', 4.5))
                iv = config['Graph'].get('interval', '1s')
                self.interval_var.set(iv)
                
                # Update the actual internal timer to match the loaded setting
                if iv in INTERVAL_OPTIONS:
                    self.current_interval_sec = INTERVAL_OPTIONS[iv]

                # Let on_device_changed handle the logic for restricted intervals
            for i in range(MAX_CHANNELS):
                sect = f'Channel_{i}'
                if sect in config:
                    self.ch_vars['active'][i].set(config[sect].getboolean('active', False))
                    self.ch_vars['factor'][i].set(config[sect].getfloat('factor', 1.0))
                    self.ch_vars['offset'][i].set(config[sect].getfloat('offset', 0.0))
        except Exception as e:
            print(f"Error parsing settings: {e}")

    def save_settings(self):
        config = configparser.ConfigParser()
        config['Graph'] = {
            'window_size': str(self.window_size_var.get()),
            'y_min': str(self.y_min.get()),
            'y_max': str(self.y_max.get()),
            'interval': self.interval_var.get()
        }
        for i in range(MAX_CHANNELS):
            config[f'Channel_{i}'] = {
                'active': str(self.ch_vars['active'][i].get()),
                'factor': str(self.ch_vars['factor'][i].get()),
                'offset': str(self.ch_vars['offset'][i].get())
            }
        with open(INI_FILE, 'w') as f: config.write(f)

    def on_close(self):
        self.is_running = False
        self.log_queue.put(None)
        self.toolbar_ref.app = None # Break circular ref
        if self.data_source: self.data_source.disconnect()
        if self.save_on_exit_var.get(): self.save_settings()
        
        with self.file_lock:
             if self.log_file and not self.log_file.closed:
                 self.log_file.close()
        self.root.destroy()

    def init_plot(self) -> list:
        return self.lines

    def _capture_data(self) -> None:
        """Read sensor data, append to buffers, and queue for logging."""
        now_ts = time.time()
        time_diff = now_ts - self.last_capture_time

        if not (self.connected and self.is_running and time_diff >= self.current_interval_sec):
            return

        raw_data = self.data_source.get_data()

        # Insert NaN gap marker if time gap exceeds threshold
        gap_threshold = self.current_interval_sec * GAP_MULTIPLIER
        if self.last_capture_time != 0 and time_diff > gap_threshold:
            pad_ts = now_ts - (self.current_interval_sec / 2)
            self.timestamps.append(pad_ts)
            self.datetime_cache.append(datetime.fromtimestamp(pad_ts))
            for i in range(MAX_CHANNELS):
                self.channel_data[i].append(float('nan'))

        self.last_capture_time = now_ts
        self.timestamps.append(now_ts)
        self.datetime_cache.append(datetime.fromtimestamp(now_ts))

        row: list = [now_ts, datetime.fromtimestamp(now_ts).strftime("%Y-%m-%d %H:%M:%S.%f")]
        for i in range(MAX_CHANNELS):
            val = raw_data[i] if i < len(raw_data) else 0.0
            self.channel_data[i].append(val)
            row.append(val)

            f = self.ch_vars['factor'][i].get()
            o = self.ch_vars['offset'][i].get()
            calc_val = (val * f) + o
            self.ch_vars['current_val_str'][i].set(f"{calc_val:.2f}")

        if self.csv_writer:
            self.log_queue.put(row)

    def _compute_gap_blueprint(self, raw_dates: list, step: int) -> tuple[list, list]:
        """
        Detects time gaps in data and creates a 'blueprint' for inserting NaNs.
        Returns:
            final_dt: List of datetimes including gap timestamps.
            indices: List of (index, is_gap) tuples to reconstruct Y-values.
        """
        gap_threshold = timedelta(seconds=self.current_interval_sec * step * GAP_MULTIPLIER)
        expanded_indices = []
        final_dt = []
        
        if raw_dates:
            expanded_indices.append((0, False))
            for i in range(1, len(raw_dates)):
                diff = raw_dates[i] - raw_dates[i-1]
                if diff > gap_threshold:
                    expanded_indices.append((i, True))  # Mark gap
                expanded_indices.append((i, False))  # Mark data

            for idx, is_gap in expanded_indices:
                if is_gap:
                    final_dt.append(raw_dates[idx-1]) # Use prev time for gap start
                else:
                    final_dt.append(raw_dates[idx])
                    
        return final_dt, expanded_indices

    def _render_plot(self) -> list:
        """Slice, downsample, and render data to plot lines. Returns line artists."""
        if not self.timestamps:
            return self.lines

        # Lazy-initialize datetime axis locator on first data arrival
        if not self._datetime_axis_initialized:
            self.ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
            self.ax.xaxis.set_major_locator(mdates.AutoDateLocator())
            self.ax.figure.canvas.draw_idle()  # Trigger redraw to apply locator
            self._datetime_axis_initialized = True

        total_points = len(self.timestamps)
        render_start = 0
        render_end = total_points
        view_min_dt = None

        if self.auto_scroll.get():
            # AUTO SCROLL MODE: Show data from the past N seconds (from slider)
            window_seconds = self.slider_to_seconds(self.window_size_var.get())
            
            if self.timestamps:
                latest_ts = self.timestamps[-1]
                cutoff_ts = latest_ts - window_seconds
                view_min_dt = datetime.fromtimestamp(cutoff_ts)
                
                # Use cached list if size hasn't changed
                current_size = len(self.timestamps)
                if current_size != self._timestamps_cache_size:
                    self._timestamps_list_cache = list(self.timestamps)
                    self._timestamps_cache_size = current_size
                
                ts_list = self._timestamps_list_cache
                render_start = max(0, bisect.bisect_left(ts_list, cutoff_ts))
                render_end = total_points
            else:
                render_end = total_points
        else:
            # MANUAL PAN/ZOOM MODE: Calculate visible points dynamically
            try:
                # Get current Axis Limits
                x_min, x_max = self.ax.get_xlim()
                
                view_min_dt = mdates.num2date(x_min).replace(tzinfo=None)
                view_max_dt = mdates.num2date(x_max).replace(tzinfo=None)
                view_min_ts = view_min_dt.timestamp()
                view_max_ts = view_max_dt.timestamp()

                # Use cached list if size hasn't changed (avoid O(N) conversion every frame)
                current_size = len(self.timestamps)
                if current_size != self._timestamps_cache_size:
                    self._timestamps_list_cache = list(self.timestamps)
                    self._timestamps_cache_size = current_size
                
                ts_list = self._timestamps_list_cache
                idx_min = bisect.bisect_left(ts_list, view_min_ts)
                idx_max = bisect.bisect_right(ts_list, view_max_ts)
                
                # Define Render Range with Buffer
                points_in_view = idx_max - idx_min
                buffer = max(points_in_view, 200)
                
                render_start = max(0, idx_min - buffer)
                render_end = min(total_points, idx_max + buffer)
                
            except Exception:
                render_start = 0
                render_end = total_points

        # --- STEP CALCULATION (RESOLUTION) ---
        points_to_render = render_end - render_start
        step = 1
        
        # In Pan mode, allow higher point count before downsampling
        target_limit = TARGET_PLOT_POINTS
        if not self.auto_scroll.get():
            target_limit = TARGET_PLOT_POINTS * 3
        
        if points_to_render > target_limit:
            step = points_to_render // target_limit
        
        step = max(1, step)

        # --- SLICING & DRAWING ---
        slice_dt_raw = list(itertools.islice(self.datetime_cache, render_start, render_end, step))
        
        # Helper Call to generate blueprint
        final_dt, expanded_indices = self._compute_gap_blueprint(slice_dt_raw, step)
        
        for i, line in enumerate(self.lines):
            if self.ch_vars['active'][i].get():
                raw_slice = list(itertools.islice(self.channel_data[i], render_start, render_end, step))
                
                f = self.ch_vars['factor'][i].get()
                o = self.ch_vars['offset'][i].get()
                
                # Reconstruct with gaps using blueprint
                calc_slice: list[float] = []
                for idx, is_gap in expanded_indices:
                    if is_gap:
                        calc_slice.append(float('nan'))
                    else:
                        if idx < len(raw_slice):
                            val = raw_slice[idx]
                            calc_slice.append((val * f) + o)
                        else:
                            calc_slice.append(float('nan'))
                            
                line.set_data(final_dt, calc_slice)
                line.set_visible(True)
            else:
                line.set_visible(False)
        
        # Handle X-Axis Limits (Only if Auto-Scrolling)
        if self.auto_scroll.get() and final_dt:
            # Use actual latest timestamp (not downsampled) to avoid edge artifacts
            actual_latest_dt = self.datetime_cache[-1] if self.datetime_cache else final_dt[-1]
            
            # Use calculated view start (smooth) if available, otherwise snap to data (old behavior)
            start_limit = view_min_dt if view_min_dt else final_dt[0]
            self.ax.set_xlim(start_limit, actual_latest_dt)
        
        return self.lines

    def update_plot(self, frame) -> list:
        """Animation callback: capture data and render plot."""
        try:
            self.check_rollover()
            self._capture_data()
            return self._render_plot()
        except Exception as e:
            # print(f"Plot Error: {e}")
            return self.lines

    def apply_scale(self) -> None:
        """Apply Y-axis limits from UI variables."""
        try:
            self.ax.set_ylim(self.y_min.get(), self.y_max.get())
            self.canvas.draw_idle()
        except ValueError:
            pass

    def toggle_capture(self) -> None:
        """Toggle data capture on/off."""
        self.is_running = not self.is_running

if __name__ == "__main__":
    root = tk.Tk()
    app = SensorApp(root)
    root.mainloop()